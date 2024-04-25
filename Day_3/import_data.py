from openpyxl import load_workbook
import mysql.connector

workbook = load_workbook('2404.xlsx')
sheet = workbook.active

values = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    print(row)
    values.append(row)

db = mysql.connector.connect(
    host = 'localhost',
    port = 3306,
    user = 'root',
    password = '12345678',
    database = 'golf_want_to_buy'
)

cursor = db.cursor()
sql = '''
    INSERT INTO products (title, price, is_necessary)
    VALUES (%s, %s, %s);
'''

cursor.executemany(sql, values)
db.commit()
print('เพิ่มข้อมูลจำนวน ' + str(cursor.rowcount) + 'แถว')

