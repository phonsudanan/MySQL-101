from openpyxl import load_workbook
import mysql.connector

#Excel
workbook = load_workbook('Import_Data2.xlsx')
sheet = workbook.active

#Database
db = mysql.connector.connect(
    host = 'localhost',
    port = 3306,
    user = 'root',
    password = '12345678',
    database = 'golf_want_to_buy'
)

cursor = db.cursor()

#โหลดข้อมูลประเภทสินค้าทั้งหมด
sql_select_categories = '''
    SELECT * 
    FROM categories
'''
cursor.execute(sql_select_categories)
categories = cursor.fetchall()
for c in categories:
    print(c)

#เปรียบเทียบข้อมูลประเภทสินค้า ถ้ายังไม่มี ให้เพิ่มข้อมูลลงไป
categories_values = []
for row in sheet.iter.rows(minrow=2, values_only=True):
    is_new = True
    category = row[3]

    for c in categories: #มาจาก Database
        if category == c[1]: #เปรียบเทียบจาก excel กับ Database
            is_new = False
            break

    if is_new:
        print((category, ))
        categories_values.append((category, ))

if len(categories_values) > 0:
    sql_insert_categories = '''
        INSERT INTO categories (title)
        VALUES (%s)
    '''
    cursor.executemany(sql_insert_categories, categories_values)
    db.commit()
    print('เพิ่มประเภทสินค้า จำนวน :' + str(cursor.rowcount) + 'แถว')
else:
    print('ไม่มีประเภทสินค้าใหม่')

#โหลดข้อมูลประเภทสินค้าทั้งหมดอีกครั้ง หลังเพิ่มข้อมูล
cursor.execute(sql_select_categories)
categories = cursor.fetchall()

#เชื่อมต่อ category_id กับสินค้าใหม่แล้วเพิ่มลงไป
products_values = []
for row in sheet.iter_rows(minrow=2, values_only=True):
    category_title = row[3]
    category_id = 'NULL'

    for c in categories:
        if category_title == c[1]:
            category_id = c[0]
            break
    product = (row[0], row[1], row[2], category_id)
    print(product)
    products_values.append(product)


sql_insert_products = '''
        INSERT INTO products (title, price, is_necessary, category_id)
        VALUES (%s, %s, %s, %s)
    '''
cursor.executemany(sql_insert_products, products_values)
db.commit()
print('เพิ่มสินค้า จำนวน :' + str(cursor.rowcount) + 'แถว')

#ปิดการเชื่อมต่อ
cursor.close()
db.close()
